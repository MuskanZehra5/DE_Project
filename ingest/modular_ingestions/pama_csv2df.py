# clean_pama_excel_openpyxl.py

import re
import pandas as pd
from pathlib import Path
import openpyxl

# ───── CONFIG ────────────────────────────────────────────────────────────────
EXCEL_FILE  = Path(r"D:\LUMS\Spring 2025\DE\DE-final-project\five-years-10-1.xlsx")
HEADER_ROW  = 2     # zero-based index: row 3 has your column labels
SKIP_FOOTER = 0     # how many rows at the bottom to drop
MIN_YEAR    = 2015  # include sheets whose start-year ≥ this
# ────────────────────────────────────────────────────────────────────────────────

def sheet_to_df(ws):
    """
    Convert an openpyxl worksheet into a pandas DataFrame,
    skipping HEADER_ROW rows, using the next row as header, then
    reading until SKIP_FOOTER rows from the end.
    """
    rows = list(ws.values)  # each row is a tuple of cell values

    # 1) the header labels:
    header = rows[HEADER_ROW]
    # 2) the data rows (drop HEADER_ROW+1 header, then skip footer)
    data_rows = rows[HEADER_ROW+1 : len(rows)-SKIP_FOOTER if SKIP_FOOTER else None]

    # 3) build DataFrame
    df = pd.DataFrame(data_rows, columns=header)

    # 4) drop any entirely empty columns/rows
    df = df.dropna(how="all").dropna(axis=1, how="all")
    return df

def load_and_clean():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    pattern = re.compile(r"^(\d{4})[-–]\d{2,4}$")

    dfs = []
    for sheet_name in wb.sheetnames:
        m = pattern.match(sheet_name)
        if not m:
            continue
        start_year = int(m.group(1))
        if start_year < MIN_YEAR:
            continue

        ws = wb[sheet_name]
        df = sheet_to_df(ws)
        df.insert(0, "Year", start_year)
        dfs.append(df)

    if not dfs:
        raise RuntimeError(f"No sheets ≥ {MIN_YEAR} found; available: {wb.sheetnames}")

    full = pd.concat(dfs, ignore_index=True)
    return full

if __name__ == "__main__":
    df = load_and_clean()
    print(f"Loaded DataFrame: {df.shape[0]} rows × {df.shape[1]} cols")
    print(df.head(10))

    out_csv = EXCEL_FILE.with_name("pama_2015_to_present.csv")
    df.to_csv(out_csv, index=False)
    print("✅ Saved cleaned CSV to:", out_csv)
